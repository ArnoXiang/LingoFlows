from flask import Flask, request, jsonify, send_from_directory, after_this_request, send_file, make_response
from flask_cors import CORS
import pymysql
import os
import uuid
from datetime import datetime, timedelta, date
import jwt
from src.api.chatGPT import get_chat_response
import re
import tempfile
import zipfile
import shutil
from flask import send_file
import logging
import json
import traceback
import time
from werkzeug.utils import secure_filename
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 配置详细的日志记录
log_dir = 'logs'
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

# 配置日志记录器
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(log_dir, 'flask_app.log')),
        logging.StreamHandler()  # 同时输出到控制台
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 配置上传文件存储路径
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# JWT配置
app.config['SECRET_KEY'] = 'lingoflows_secret_key'
app.config['JWT_EXPIRATION_DELTA'] = timedelta(days=1)

# 配置CORS，允许所有来源，包括认证请求
CORS(app, resources={r"/api/*": {"origins": "*"}}, 
     supports_credentials=True, 
     allow_headers=["Content-Type", "Authorization", "Access-Control-Allow-Credentials"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])

# 配置 MySQL 数据库连接
db = pymysql.connect(
    host='localhost',
    user='root',  
    password='root',  
    database='l10n_management',
    cursorclass=pymysql.cursors.DictCursor
)

# 定义JSON序列化帮助函数
def json_serial(obj):
    """JSON序列化器，处理不可序列化的对象如datetime"""
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError (f"Type {type(obj)} not serializable")

# 用户认证相关接口
@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    print(f"尝试登录 - 用户名: {username}")
    
    if not username or not password:
        print("错误: 用户名和密码是必需的")
        return jsonify({"error": "Username and password are required"}), 400
    
    with db.cursor() as cur:
        # 查询用户
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cur.fetchone()
        
        if not user or user['password'] != password:  
            print("错误: 无效的用户名或密码")
            return jsonify({"error": "Invalid username or password"}), 401
        
        print(f"登录成功 - 用户ID: {user['id']}, 角色: {user['role']}")
        
        # 生成JWT令牌
        payload = {
            'id': user['id'],  # 使用'id'作为键名，与其他地方保持一致
            'user_id': user['id'],  # 同时保留'user_id'键名以兼容可能的旧代码
            'username': user['username'],
            'role': user['role'],
            'exp': datetime.utcnow() + app.config['JWT_EXPIRATION_DELTA']
        }
        token = jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')
        
        print(f"生成的令牌负载: {payload}")
        
        return jsonify({
            "token": token,
            "user": {
                "id": user['id'],
                "username": user['username'],
                "role": user['role'],
                "name": user['name']
            }
        })


@app.route('/api/users/current', methods=['GET'])
def get_current_user():
    auth_header = request.headers.get('Authorization')
    if not auth_header:
        return jsonify({"error": "Authorization header is missing"}), 401
    
    try:
        token = auth_header.split(" ")[1]
        payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
        
        with db.cursor() as cur:
            cur.execute("SELECT id, username, role, name FROM users WHERE id = %s", (payload['user_id'],))
            user = cur.fetchone()
            
            if not user:
                return jsonify({"error": "User not found"}), 404
            
            return jsonify(user)
    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token has expired"}), 401
    except (jwt.InvalidTokenError, IndexError):
        return jsonify({"error": "Invalid token"}), 401

# 验证令牌的装饰器
def token_required(f):
    def decorated(*args, **kwargs):
        print("------- 开始验证令牌 -------")
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            print("错误: 请求中缺少Authorization头")
            print(f"请求头: {dict(request.headers)}")
            return jsonify({"error": "Authorization header is missing"}), 401
        
        print(f"收到Authorization头: {auth_header}")
    
        try:
            # 检查格式是否符合"Bearer <token>"
            if not auth_header.startswith('Bearer '):
                print("错误: Authorization头格式错误，应该以'Bearer '开头")
                return jsonify({"error": "Invalid authorization format, expected 'Bearer <token>'"}), 401
            
            token = auth_header.split(" ")[1]
            print(f"提取的令牌: {token[:10]}...")
            
            try:
                payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
                print(f"令牌解码成功，负载: {payload}")

                
                # 检查必要的字段
                if 'user_id' not in payload:
                    print("错误: 令牌负载中缺少user_id字段")
                    return jsonify({"error": "Token payload is missing user_id field"}), 401
                
                # 查询用户是否存在
                with db.cursor() as cur:
                    cur.execute("SELECT id, username, role FROM users WHERE id = %s", (payload['user_id'],))
                    user = cur.fetchone()
                    
                    if not user:
                        print(f"错误: 令牌中的用户ID {payload['user_id']} 不存在")
                        return jsonify({"error": "User not found"}), 401
                    
                    print(f"验证用户存在: {user['username']}, 角色: {user['role']}")
                
                request.user = payload
                print("令牌验证成功")
                return f(*args, **kwargs)


            except jwt.ExpiredSignatureError:
                print("错误: 令牌已过期")
                return jsonify({"error": "Token has expired"}), 401
            except jwt.InvalidTokenError as e:
                print(f"错误: 无效的令牌 - {str(e)}")
                return jsonify({"error": f"Invalid token: {str(e)}"}), 401
        except Exception as e:
            print(f"验证令牌时发生未预期的错误: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": "Authentication error"}), 500
    
    decorated.__name__ = f.__name__
    return decorated
        
# 项目相关接口
@app.route('/api/projects', methods=['GET'])
@token_required
def get_projects():
    user_role = request.user.get('role')
    # 尝试从不同的键名获取用户ID
    user_id = request.user.get('user_id') or request.user.get('id')
    
    # 检查是否有特殊的查询参数
    role_param = request.args.get('role')
    all_param = request.args.get('all')
    fallback_param = request.args.get('fallback')
    
    print(f"获取项目 - 用户角色: {user_role}, 用户ID: {user_id}")
    print(f"查询参数: role={role_param}, all={all_param}, fallback={fallback_param}")
    print(f"完整的用户信息: {request.user}")
    
    with db.cursor() as cur:
        # 如果是LM或FT角色，或者请求包含all=true参数，允许查看所有项目
        if user_role in ['LM', 'FT'] or all_param == 'true' or fallback_param == 'true':
            # LM/FT可以查看所有项目
            print(f"{user_role}用户或特殊查询 - 查询所有项目")
            # 修改SQL，添加关联查询以获取请求者名称
            cur.execute("""
                SELECT p.*, u.name as requesterName  
                FROM projectname p
                LEFT JOIN users u ON p.created_by = u.id
            """)
        else:
            # BO只能查看自己提交的项目
            print(f"非LM/FT用户 - 只查询用户ID为 {user_id} 的项目")
            # 修改SQL，添加关联查询以获取请求者名称
            cur.execute("""
                SELECT p.*, u.name as requesterName  
                FROM projectname p
                LEFT JOIN users u ON p.created_by = u.id
                WHERE p.created_by = %s
            """, (user_id,))
        
        projects = cur.fetchall()
        print(f"查询结果: 找到 {len(projects)} 个项目")
        
        # 如果是LM/FT用户但没有找到项目，检查是否有项目的created_by为NULL
        if user_role in ['LM', 'FT'] and len(projects) == 0:
            print(f"{user_role}用户没有找到项目，检查是否有项目的created_by为NULL")
            cur.execute("SELECT COUNT(*) as count FROM projectname WHERE created_by IS NULL")
            null_count = cur.fetchone()['count']
            print(f"created_by为NULL的项目数: {null_count}")
            
            if null_count > 0:
                print("更新created_by为NULL的项目，设置为当前用户ID")
                cur.execute("UPDATE projectname SET created_by = %s WHERE created_by IS NULL", (user_id,))
                db.commit()
                
                # 重新查询所有项目，包括请求者名称
                cur.execute("""
                    SELECT p.*, u.name as requesterName  
                    FROM projectname p
                    LEFT JOIN users u ON p.created_by = u.id
                """)
                projects = cur.fetchall()
                print(f"更新后查询结果: 找到 {len(projects)} 个项目")
    
    return jsonify(projects)

@app.route('/api/projects/<int:project_id>', methods=['GET'])
@token_required
def get_project(project_id):
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    # 添加请求者名称查询
    with db.cursor() as cur:
        cur.execute("""
            SELECT p.*, u.name as requesterName 
            FROM projectname p
            LEFT JOIN users u ON p.created_by = u.id
            WHERE p.id = %s
        """, (project_id,))
        project = cur.fetchone()
        
        if not project:
            return jsonify({"error": "Project not found"}), 404
            
        # 非LM用户只能查看自己的项目
        if user_role != 'LM' and project['created_by'] != user_id:
            return jsonify({"error": "You don't have permission to access this project"}), 403
            
        return jsonify(project)

@app.route('/api/projects', methods=['POST'])
@token_required
def create_project():
    data = request.json
    user_id = request.user.get('user_id')
    
    # 从请求中提取项目数据
    project_name = data.get('projectName')
    request_name = data.get('requestName')
    project_manager = data.get('projectManager', 'Yizhuo Xiang')  
    project_status = data.get('projectStatus', 'pending')  
    
    # 其他项目数据
    source_language = data.get('sourceLanguage')
    target_languages = ','.join(data.get('targetLanguages', []))
    word_count = data.get('wordCount', 0)
    expected_delivery_date = data.get('expectedDeliveryDate')
    additional_requirements = ','.join(data.get('additionalRequirements', []))
    
    # 创建项目记录
    with db.cursor() as cur:
        sql = """
        INSERT INTO projectname (
            projectName, projectStatus, requestName, projectManager, createTime,
            sourceLanguage, targetLanguages, wordCount, expectedDeliveryDate, additionalRequirements,
            taskTranslation, taskLQA, taskTranslationUpdate, taskLQAReportFinalization, created_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cur.execute(sql, (
            project_name, project_status, request_name, project_manager, datetime.now(),
            source_language, target_languages, word_count, expected_delivery_date, additional_requirements,
            'not_started', 'not_started', 'not_started', 'not_started', user_id
        ))
        db.commit()
        project_id = cur.lastrowid
    
    return jsonify({"id": project_id, "message": "Project created successfully"}), 201

@app.route('/api/projects/<int:project_id>', methods=['PUT'])
@token_required
def update_project(project_id):
    data = request.json
    user_role = request.user.get('role')
    user_id = request.user.get('user_id') or request.user.get('id')
    
    print(f"更新项目 - 项目ID: {project_id}, 用户角色: {user_role}, 用户ID: {user_id}")
    print(f"更新数据: {data}")
    
    # 检查用户是否有权限更新此项目
    with db.cursor() as cur:
        if user_role in ['LM', 'FT']:
            # LM和FT可以更新任何项目
            cur.execute("SELECT id FROM projectname WHERE id = %s", (project_id,))
        else:
            # BO只能更新自己的项目
            cur.execute("SELECT id FROM projectname WHERE id = %s AND created_by = %s", (project_id, user_id))
        
        project = cur.fetchone()
        if not project:
            return jsonify({"error": "Project not found or you don't have permission"}), 404
    
    try:
        # 从请求中提取项目数据
        update_fields = []
        update_values = []
        
        # 检查并添加更新字段 - 基本信息
        if 'projectName' in data:
            update_fields.append("projectName = %s")
            update_values.append(data['projectName'])
        
        if 'projectStatus' in data:
            update_fields.append("projectStatus = %s")
            update_values.append(data['projectStatus'])
        
        if 'requestName' in data:
            update_fields.append("requestName = %s")
            update_values.append(data['requestName'])
        
        if 'projectManager' in data:
            update_fields.append("projectManager = %s")
            update_values.append(data['projectManager'])
        
        # 检查并添加更新字段 - 任务状态
        if 'taskTranslation' in data:
            update_fields.append("taskTranslation = %s")
            update_values.append(data['taskTranslation'])
        
        if 'taskLQA' in data:
            update_fields.append("taskLQA = %s")
            update_values.append(data['taskLQA'])
        
        if 'taskTranslationUpdate' in data:
            update_fields.append("taskTranslationUpdate = %s")
            update_values.append(data['taskTranslationUpdate'])
        
        if 'taskLQAReportFinalization' in data:
            update_fields.append("taskLQAReportFinalization = %s")
            update_values.append(data['taskLQAReportFinalization'])
        
        # 检查并添加更新字段 - 任务详细信息
        if 'translationAssignee' in data:
            update_fields.append("translationAssignee = %s")
            update_values.append(data['translationAssignee'])
        
        if 'translationDeadline' in data:
            update_fields.append("translationDeadline = %s")
            try:
                if data['translationDeadline']:
                    if isinstance(data['translationDeadline'], str):
                        date_str = data['translationDeadline'].split('T')[0] if 'T' in data['translationDeadline'] else data['translationDeadline']
                        update_values.append(date_str)
                    else:
                        update_values.append(data['translationDeadline'])
                else:
                    update_values.append(None)
            except Exception as e:
                print(f"处理翻译截止日期时出错: {e}")
                update_values.append(None)
        
        if 'translationNotes' in data:
            update_fields.append("translationNotes = %s")
            update_values.append(data['translationNotes'])
        
        if 'lqaAssignee' in data:
            update_fields.append("lqaAssignee = %s")
            update_values.append(data['lqaAssignee'])
        
        if 'lqaDeadline' in data:
            update_fields.append("lqaDeadline = %s")
            try:
                if data['lqaDeadline']:
                    if isinstance(data['lqaDeadline'], str):
                        date_str = data['lqaDeadline'].split('T')[0] if 'T' in data['lqaDeadline'] else data['lqaDeadline']
                        update_values.append(date_str)
                    else:
                        update_values.append(data['lqaDeadline'])
                else:
                    update_values.append(None)
            except Exception as e:
                print(f"处理LQA截止日期时出错: {e}")
                update_values.append(None)
        
        if 'lqaNotes' in data:
            update_fields.append("lqaNotes = %s")
            update_values.append(data['lqaNotes'])
        
        # 添加翻译更新和LQA报告定稿的字段
        if 'translationUpdateAssignee' in data:
            update_fields.append("translationUpdateAssignee = %s")
            update_values.append(data['translationUpdateAssignee'])
        
        if 'translationUpdateDeadline' in data:
            update_fields.append("translationUpdateDeadline = %s")
            try:
                if data['translationUpdateDeadline']:
                    if isinstance(data['translationUpdateDeadline'], str):
                        date_str = data['translationUpdateDeadline'].split('T')[0] if 'T' in data['translationUpdateDeadline'] else data['translationUpdateDeadline']
                        update_values.append(date_str)
                    else:
                        update_values.append(data['translationUpdateDeadline'])
                else:
                    update_values.append(None)
            except Exception as e:
                print(f"处理翻译更新截止日期时出错: {e}")
                update_values.append(None)
        
        if 'translationUpdateNotes' in data:
            update_fields.append("translationUpdateNotes = %s")
            update_values.append(data['translationUpdateNotes'])
        
        if 'lqaReportFinalizationAssignee' in data:
            update_fields.append("lqaReportFinalizationAssignee = %s")
            update_values.append(data['lqaReportFinalizationAssignee'])
        
        if 'lqaReportFinalizationDeadline' in data:
            update_fields.append("lqaReportFinalizationDeadline = %s")
            try:
                if data['lqaReportFinalizationDeadline']:
                    if isinstance(data['lqaReportFinalizationDeadline'], str):
                        date_str = data['lqaReportFinalizationDeadline'].split('T')[0] if 'T' in data['lqaReportFinalizationDeadline'] else data['lqaReportFinalizationDeadline']
                        update_values.append(date_str)
                    else:
                        update_values.append(data['lqaReportFinalizationDeadline'])
                else:
                    update_values.append(None)
            except Exception as e:
                print(f"处理LQA报告定稿截止日期时出错: {e}")
                update_values.append(None)
        
        if 'lqaReportFinalizationNotes' in data:
            update_fields.append("lqaReportFinalizationNotes = %s")
            update_values.append(data['lqaReportFinalizationNotes'])
        
        # 检查并添加更新字段 - 其他信息
        if 'sourceLanguage' in data:
            update_fields.append("sourceLanguage = %s")
            update_values.append(data['sourceLanguage'])
        
        if 'targetLanguages' in data:
            update_fields.append("targetLanguages = %s")
            update_values.append(data['targetLanguages'])
        
        if 'wordCount' in data:
            update_fields.append("wordCount = %s")
            update_values.append(data['wordCount'])
        
        if 'expectedDeliveryDate' in data:
            update_fields.append("expectedDeliveryDate = %s")

            try:
                if data['expectedDeliveryDate']:
                    # 如果是日期对象的字符串表示，尝试转换为日期对象
                    if isinstance(data['expectedDeliveryDate'], str):
                        # 移除可能的时区信息
                        date_str = data['expectedDeliveryDate'].split('T')[0] if 'T' in data['expectedDeliveryDate'] else data['expectedDeliveryDate']
                        update_values.append(date_str)
                    else:
                        update_values.append(data['expectedDeliveryDate'])
                else:
                    update_values.append(None)
            except Exception as e:
                print(f"处理日期时出错: {e}")
                update_values.append(None)
        
        if 'additionalRequirements' in data:
            update_fields.append("additionalRequirements = %s")
            update_values.append(data['additionalRequirements'])
        
        # 如果有需要更新的项目字段，执行更新
        project_updated = False
        if update_fields:
            # 构建更新SQL
            sql = "UPDATE projectname SET " + ", ".join(update_fields) + " WHERE id = %s"
            update_values.append(project_id)
            
            print(f"更新SQL: {sql}")
            print(f"更新值: {update_values}")
            
            # 执行更新
            with db.cursor() as cur:
                # 首先检查表是否有必要的字段
                try:
                    cur.execute("SHOW COLUMNS FROM projectname LIKE 'translationAssignee'")
                    has_translation_assignee = cur.fetchone() is not None
                    
                    if not has_translation_assignee:
                        # 添加必要的字段
                        print("添加任务详细信息字段到projectname表")
                        cur.execute("""
                            ALTER TABLE projectname 
                            ADD COLUMN translationAssignee VARCHAR(100),
                            ADD COLUMN translationDeadline DATE,
                            ADD COLUMN translationNotes TEXT,
                            ADD COLUMN lqaAssignee VARCHAR(100),
                            ADD COLUMN lqaDeadline DATE,
                            ADD COLUMN lqaNotes TEXT
                        """)
                        db.commit()
                        print("成功添加字段")
                    
                    # 检查是否有翻译更新和LQA报告定稿的字段
                    cur.execute("SHOW COLUMNS FROM projectname LIKE 'translationUpdateAssignee'")
                    has_translation_update_assignee = cur.fetchone() is not None
                    
                    if not has_translation_update_assignee:
                        # 添加必要的字段
                        print("添加翻译更新和LQA报告定稿字段到projectname表")
                        cur.execute("""
                            ALTER TABLE projectname 
                            ADD COLUMN translationUpdateAssignee VARCHAR(100),
                            ADD COLUMN translationUpdateDeadline DATE,
                            ADD COLUMN translationUpdateNotes TEXT,
                            ADD COLUMN lqaReportFinalizationAssignee VARCHAR(100),
                            ADD COLUMN lqaReportFinalizationDeadline DATE,
                            ADD COLUMN lqaReportFinalizationNotes TEXT
                        """)
                        db.commit()
                        print("成功添加翻译更新和LQA报告定稿字段")
                except Exception as e:
                    print(f"检查或添加字段时出错: {e}")
                    # 继续执行，即使添加字段失败
                
                # 执行更新
                cur.execute(sql, update_values)
                db.commit()
                project_updated = cur.rowcount > 0
                print(f"项目基本信息更新结果: {'成功' if project_updated else '无变化'}, 影响行数: {cur.rowcount}")
                
                # 已移除对rowcount为0的检查，继续处理任务分配数据
        else:
            print("没有项目基本信息需要更新")
            
        # 处理任务分配数据
        task_assignments_updated = False
        if 'taskAssignments' in data and isinstance(data['taskAssignments'], list) and data['taskAssignments']:
            try:
                # 检查表是否存在
                with db.cursor() as cur:
                    cur.execute("SHOW TABLES LIKE 'project_task_assignments'")
                    table_exists = cur.fetchone() is not None
                    
                    if not table_exists:
                        # 创建项目任务分配表
                        print("创建项目任务分配表")
                        cur.execute("""
                            CREATE TABLE IF NOT EXISTS project_task_assignments (
                              id INT AUTO_INCREMENT PRIMARY KEY,
                              project_id INT NOT NULL,
                              task_type ENUM('translation', 'lqa', 'translationUpdate', 'lqaReportFinalization') NOT NULL,
                              language VARCHAR(10) NOT NULL,
                              assignee VARCHAR(100) NOT NULL,
                              deadline DATE,
                              notes TEXT,
                              createTime DATETIME NOT NULL,
                              created_by INT NOT NULL,
                              FOREIGN KEY (project_id) REFERENCES projectname(id) ON DELETE CASCADE,
                              FOREIGN KEY (created_by) REFERENCES users(id)
                            )
                        """)
                        
                        # 创建索引
                        cur.execute("CREATE INDEX idx_task_assignments_project ON project_task_assignments(project_id)")
                        cur.execute("CREATE INDEX idx_task_assignments_task ON project_task_assignments(task_type)")
                        cur.execute("CREATE INDEX idx_task_assignments_language ON project_task_assignments(language)")
                        
                        db.commit()
                        print("成功创建项目任务分配表")
                    
                    # 查询当前项目的所有任务分配
                    cur.execute("SELECT COUNT(*) as count FROM project_task_assignments WHERE project_id = %s", (project_id,))
                    result = cur.fetchone()
                    existing_assignments_count = result['count'] if result else 0
                    
                    # 先删除该项目的所有任务分配
                    cur.execute("DELETE FROM project_task_assignments WHERE project_id = %s", (project_id,))
                    deleted_count = cur.rowcount
                    print(f"已删除 {deleted_count} 个现有任务分配")
                    
                    # 添加新的任务分配
                    insert_count = 0
                    for assignment in data['taskAssignments']:
                        cur.execute("""
                            INSERT INTO project_task_assignments 
                            (project_id, task_type, language, assignee, deadline, notes, createTime, created_by)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            project_id,
                            assignment['task_type'],
                            assignment['language'],
                            assignment['assignee'],
                            assignment['deadline'],
                            assignment['notes'],
                            datetime.now(),
                            user_id
                        ))
                        insert_count += 1
                    
                    db.commit()
                    print(f"成功保存 {insert_count} 个任务分配")
                    
                    # 判断任务分配是否有实际变化
                    task_assignments_updated = insert_count != existing_assignments_count or insert_count > 0
                    print(f"任务分配更新结果: {'成功' if task_assignments_updated else '无变化'}")
            except Exception as e:
                print(f"保存任务分配时出错: {e}")
                db.rollback()
                # 如果只有任务分配保存失败，并且项目更新成功，应该返回部分成功的消息
                if project_updated:
                    return jsonify({
                        "message": "Project updated but task assignments failed", 
                        "error": str(e)
                    }), 206  # 206 Partial Content
                return jsonify({"error": f"Failed to update task assignments: {str(e)}"}), 500
        
        # 如果项目基本信息或任务分配有更新，返回成功
        if project_updated or task_assignments_updated:
            return jsonify({"message": "Project updated successfully"})
        else:
            # 如果既没有项目字段更新也没有任务分配更新，返回提示但不是错误
            return jsonify({"message": "No changes detected"}), 200
    except Exception as e:
        print(f"更新项目时出错: {e}")
        db.rollback()
        return jsonify({"error": f"Failed to update project: {str(e)}"}), 500

@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
@token_required
def delete_project(project_id):
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    # 只有LM可以删除项目
    if user_role != 'LM':
        return jsonify({"error": "Only LM can delete projects"}), 403
    
    with db.cursor() as cur:
        cur.execute("DELETE FROM projectname WHERE id = %s", (project_id,))
        db.commit()
        if cur.rowcount == 0:
            return jsonify({"error": "Project not found"}), 404
    
    return jsonify({"message": "Project deleted successfully"})

# 请求相关接口
@app.route('/api/requests', methods=['GET'])
@token_required
def get_requests():
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    with db.cursor() as cur:
        if user_role == 'LM':
            # LM可以查看所有请求
            cur.execute("SELECT * FROM requests")
        else:
            # BO只能查看自己的请求
            cur.execute("SELECT * FROM requests WHERE created_by = %s", (user_id,))
        
        requests = cur.fetchall()
    return jsonify(requests)

@app.route('/api/requests', methods=['POST'])
@token_required
def create_request():
    data = request.json
    user_id = request.user.get('user_id')
    
    # 使用ASCII=False确保中文正常显示
    logger.info(f"收到创建请求: {json.dumps(data, ensure_ascii=False, default=json_serial)}")
    
    # 从请求中提取数据
    request_name = data.get('requestName')
    request_background = data.get('requestBackground')
    source_language = data.get('sourceLanguage')
    target_languages = ','.join(data.get('targetLanguages', []))
    word_count = data.get('wordCount', 0)
    additional_requirements = ','.join(data.get('additionalRequirements', []))
    expected_delivery_date = data.get('expectedDeliveryDate')
    files = ','.join(data.get('files', []))
    # 提取文件ID列表 - 这个是新增的
    file_ids = data.get('fileIds', [])
    
    # 增强日志记录，检查fileIds的内容和类型
    logger.info(f"提取的文件IDs: {file_ids}, 类型: {type(file_ids)}")
    
    if isinstance(file_ids, list):
        logger.info(f"文件ID列表长度: {len(file_ids)}")
        for i, file_id in enumerate(file_ids):
            logger.info(f"文件ID[{i}]: {file_id}, 类型: {type(file_id)}")
    else:
        logger.warning(f"fileIds不是列表类型: {type(file_ids)}")
        # 尝试转换为列表
        if isinstance(file_ids, str):
            try:
                file_ids = [int(id.strip()) for id in file_ids.split(',') if id.strip()]
                logger.info(f"从字符串转换后的fileIds: {file_ids}")
            except ValueError as e:
                logger.error(f"无法将字符串转换为文件ID列表: {str(e)}")
                file_ids = []
    
    # 确保所有的文件ID都是整数类型
    original_file_ids = file_ids.copy()
    file_ids = [int(id) if isinstance(id, str) and id.isdigit() else id for id in file_ids]
    file_ids = [id for id in file_ids if isinstance(id, int)]
    
    if file_ids != original_file_ids:
        logger.info(f"文件ID经过类型转换，原始值: {original_file_ids}, 处理后: {file_ids}")
    else:
        logger.info(f"文件ID无需类型转换，保持原样: {file_ids}")
    
    # 创建请求记录
    with db.cursor() as cur:
        sql = """
        INSERT INTO requests (
            requestName, requestBackground, sourceLanguage, targetLanguages,
            wordCount, additionalRequirements, expectedDeliveryDate, files, status, createTime, created_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cur.execute(sql, (
            request_name, request_background, source_language, target_languages,
            word_count, additional_requirements, expected_delivery_date, files, 'pending', datetime.now(), user_id
        ))
        db.commit()
        request_id = cur.lastrowid
        
        logger.info(f"请求创建成功，ID: {request_id}")
        
        # 自动创建项目
        project_name = f"{request_name} 项目"
        project_manager = 'Yizhuo Xiang'  # 默认项目经理
        project_status = 'pending'  # 默认状态为待处理
        
        # 创建项目记录
        project_sql = """
        INSERT INTO projectname (
            projectName, projectStatus, requestName, projectManager, createTime,
            sourceLanguage, targetLanguages, wordCount, expectedDeliveryDate, additionalRequirements,
            taskTranslation, taskLQA, taskTranslationUpdate, taskLQAReportFinalization, created_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cur.execute(project_sql, (
            project_name, project_status, request_name, project_manager, datetime.now(),
            source_language, target_languages, word_count, expected_delivery_date, additional_requirements,
            'not_started', 'not_started', 'not_started', 'not_started', user_id
        ))
        db.commit()
        project_id = cur.lastrowid
        
        logger.info(f"项目创建成功，ID: {project_id}")
        
        # 更新请求记录，关联项目ID
        cur.execute("UPDATE requests SET project_id = %s WHERE id = %s", (project_id, request_id))
        db.commit()
        
        # 如果提供了文件ID，自动创建项目文件关联 - 这部分是新增的
        if file_ids:
            logger.info(f"开始创建项目文件关联，项目ID: {project_id}, 文件IDs: {file_ids}")
            try:
                # 确认文件是否存在且属于当前用户
                file_ids_str = ','.join([str(id) for id in file_ids])
                logger.info(f"文件ID列表字符串: {file_ids_str}")
                
                # 直接构建SQL，避免参数展开问题
                file_query = f"""
                    SELECT id FROM files 
                    WHERE id IN ({file_ids_str}) AND uploaded_by = {user_id} AND isDeleted = FALSE
                """
                logger.info(f"执行查询: {file_query}")
                
                cur.execute(file_query)
                found_files = cur.fetchall()
                
                logger.info(f"查询结果 - 找到文件: {json.dumps(found_files, default=json_serial)}")
                
                # 获取找到的文件ID列表
                found_file_ids = [f['id'] for f in found_files]
                logger.info(f"找到的文件IDs: {found_file_ids}, 数量: {len(found_file_ids)}")
                
                # 只要找到任何有效文件就创建关联
                if found_file_ids:
                    # 获取文件名，用作files字段
                    file_names_query = f"""
                        SELECT id, originalName FROM files 
                        WHERE id IN ({file_ids_str})
                    """
                    logger.info(f"查询文件名: {file_names_query}")
                    
                    cur.execute(file_names_query)
                    file_names_result = cur.fetchall()
                    file_names = [f['originalName'] for f in file_names_result]
                    files_str = ', '.join(file_names) if file_names else 'files from request'
                    
                    logger.info(f"获取到的文件名: {files_str}")
                    
                    # 检查project_files表是否有files字段
                    try:
                        cur.execute("SHOW COLUMNS FROM project_files LIKE 'files'")
                        has_files_column = cur.fetchone() is not None
                        logger.info(f"project_files表是否有files列: {has_files_column}")
                    except Exception as e:
                        logger.error(f"检查project_files表结构时出错: {str(e)}")
                        has_files_column = True  # 假设存在，避免中断流程
                    
                    # 创建项目文件记录
                    if has_files_column:
                        # 如果有files字段，正常插入
                        sql = """
                        INSERT INTO project_files (
                            projectId, fileType, notes, files, uploadTime, created_by
                        ) VALUES (%s, %s, %s, %s, %s, %s)
                        """
                        cur.execute(sql, (
                            project_id, 'source', 'Files uploaded with request submission', 
                            files_str, datetime.now(), user_id
                        ))
                    else:
                        # 如果没有files字段，忽略它
                        sql = """
                        INSERT INTO project_files (
                            projectId, fileType, notes, uploadTime, created_by
                        ) VALUES (%s, %s, %s, %s, %s)
                        """
                        cur.execute(sql, (
                            project_id, 'source', 'Files uploaded with request submission', 
                            datetime.now(), user_id
                        ))
                    
                    db.commit()
                    project_file_id = cur.lastrowid
                    
                    logger.info(f"创建项目文件记录成功，ID: {project_file_id}")
                    
                    # 创建项目文件与原始文件的关联
                    inserted_mappings = 0
                    for file_id in found_file_ids:
                        try:
                            sql = """
                            INSERT INTO project_file_mappings (
                                project_file_id, file_id
                            ) VALUES (%s, %s)
                            """
                            cur.execute(sql, (project_file_id, file_id))
                            logger.info(f"创建文件映射: 项目文件ID {project_file_id} -> 文件ID {file_id}")
                            inserted_mappings += 1
                        except Exception as mapping_err:
                            logger.error(f"创建单个文件映射时出错 (文件ID {file_id}): {str(mapping_err)}")
                    
                    if inserted_mappings > 0:
                        db.commit()
                        logger.info(f"已为请求 {request_id} 创建项目文件关联，关联了 {inserted_mappings} 个文件")
                    else:
                        logger.warning(f"尝试关联 {len(found_file_ids)} 个文件，但全部失败")
                else:
                    logger.warning(f"没有找到有效的文件 - 请求的文件IDs: {file_ids}, 用户ID: {user_id}")
                    
                    # 再次查询这些文件存在但可能不属于当前用户的情况
                    check_query = f"""
                        SELECT id, uploaded_by, isDeleted FROM files 
                        WHERE id IN ({file_ids_str})
                    """
                    logger.info(f"执行额外查询验证文件存在性: {check_query}")
                    
                    cur.execute(check_query)
                    check_results = cur.fetchall()
                    
                    if check_results:
                        logger.warning(f"找到文件但不符合条件: {json.dumps(check_results, default=json_serial)}")
                        for file in check_results:
                            if file['uploaded_by'] != user_id:
                                logger.warning(f"文件ID {file['id']} 属于用户 {file['uploaded_by']}，而不是当前用户 {user_id}")
                            if file['isDeleted']:
                                logger.warning(f"文件ID {file['id']} 已被标记为删除")
                    else:
                        logger.warning(f"找不到任何ID在 {file_ids_str} 中的文件记录")
                        
                    # 检查用户是否有权限访问这些文件
                    user_query = f"""
                        SELECT id, username, role FROM users 
                        WHERE id = {user_id}
                    """
                    logger.info(f"查询用户信息: {user_query}")
                    
                    cur.execute(user_query)
                    user_info = cur.fetchone()
                    
                    if user_info:
                        logger.info(f"用户信息: {json.dumps(user_info, default=json_serial)}")
                    else:
                        logger.warning(f"找不到ID为 {user_id} 的用户")
            except Exception as e:
                logger.error(f"创建项目文件关联时出错: {str(e)}")
                logger.error(f"文件IDs: {file_ids}")
                logger.error(f"项目ID: {project_id}")
                logger.error(f"用户ID: {user_id}")
                # 详细记录异常信息
                logger.error(traceback.format_exc())
                # 不阻止请求创建，仅记录错误
        else:
            logger.info("未提供文件IDs，跳过创建项目文件关联")
    
    return jsonify({
        "id": request_id, 
        "project_id": project_id,
        "message": "Request submitted and project created successfully"
    }), 201

# 文件上传接口
@app.route('/api/upload', methods=['POST'])
@token_required
def upload_file():
    logger.info(f"收到文件上传请求，用户: {request.user.get('username')}, ID: {request.user.get('user_id')}")
    
    # 检查请求中是否包含'file'部分
    if 'file' not in request.files:
        logger.error("错误: 请求中没有'file'部分")
        logger.debug(f"请求files字典内容: {request.files}")
        logger.debug(f"请求form字典内容: {request.form}")
        return jsonify({"error": "No file part"}), 400
    
    file = request.files['file']
    logger.info(f"获取到文件: {file.filename}, 类型: {file.content_type if hasattr(file, 'content_type') else '未知类型'}")
    
    if file.filename == '':
        logger.error("错误: 未选择文件或文件名为空")
        return jsonify({"error": "No selected file"}), 400
    
    user_id = request.user.get('user_id')
    logger.info(f"上传者ID: {user_id}")
    
    try:
        # 确保上传目录存在
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            logger.info(f"上传目录不存在，正在创建: {app.config['UPLOAD_FOLDER']}")
            os.makedirs(app.config['UPLOAD_FOLDER'])
        
        # 生成唯一文件名
        unique_filename = str(uuid.uuid4()) + '_' + file.filename
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        logger.info(f"生成的唯一文件名: {unique_filename}")
        logger.debug(f"文件将保存到: {file_path}")
        
        # 保存文件
        file.save(file_path)
        logger.info(f"文件成功保存到磁盘")
        
        # 获取文件大小和类型
        file_size = os.path.getsize(file_path)
        file_type = file.content_type if hasattr(file, 'content_type') else None
        logger.info(f"文件大小: {file_size} 字节, 类型: {file_type}")
        
        # 将文件信息存储到数据库
        with db.cursor() as cur:
            sql = """
            INSERT INTO files (
                filename, originalName, fileType, fileSize, filePath, 
                uploadTime, uploaded_by
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            cur.execute(sql, (
                unique_filename, file.filename, file_type, file_size, file_path,
                datetime.now(), user_id
            ))
            db.commit()
            file_id = cur.lastrowid
            logger.info(f"文件信息已保存到数据库，ID: {file_id}")
            
            # 确认文件记录是否成功创建
            cur.execute("SELECT * FROM files WHERE id = %s", (file_id,))
            file_record = cur.fetchone()
            logger.debug(f"验证文件记录: {json.dumps(file_record, default=json_serial) if file_record else 'Not found'}")
        
        response_data = {
            "file_id": file_id,
            "filename": unique_filename,
            "originalName": file.filename,
            "url": f"/api/files/{unique_filename}"
        }
        logger.info(f"返回上传响应: {json.dumps(response_data)}")
        return jsonify(response_data)
    except Exception as e:
        logger.error(f"文件上传处理过程中发生错误: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": f"Error processing file upload: {str(e)}"}), 500

# 文件下载接口
@app.route('/api/files/<filename>', methods=['GET'])
@token_required
def get_file(filename):
    user_id = request.user.get('user_id')
    user_role = request.user.get('role')
    
    # 检查文件是否存在并获取文件信息
    with db.cursor() as cur:
        cur.execute("SELECT * FROM files WHERE filename = %s AND isDeleted = FALSE", (filename,))
        file_info = cur.fetchone()
        
        if not file_info:
            return jsonify({"error": "File not found"}), 404
        
        # 检查权限 - LM 可以访问所有文件，BO 只能访问自己上传的或有权限的文件
        if user_role != 'LM' and file_info['uploaded_by'] != user_id:
            # 检查是否有显式权限
            cur.execute("""
                SELECT * FROM file_permissions 
                WHERE file_id = %s AND user_id = %s AND can_view = TRUE
            """, (file_info['id'], user_id))
            permission = cur.fetchone()
            
            if not permission:
                # 检查是否是项目文件，且用户属于该项目
                cur.execute("""
                    SELECT pf.* FROM project_file_mappings pfm
                    JOIN project_files pf ON pfm.project_file_id = pf.id
                    JOIN projectname p ON pf.projectId = p.id
                    WHERE pfm.file_id = %s AND p.created_by = %s
                """, (file_info['id'], user_id))
                project_file = cur.fetchone()
                
                if not project_file:
                    return jsonify({"error": "You don't have permission to access this file"}), 403
    
    # 用户有权访问文件，返回文件
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# 项目文件接口
@app.route('/api/project-files', methods=['POST'])
@token_required
def create_project_file():
    data = request.json
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    project_id = data.get('projectId')
    file_ids = data.get('fileIds', [])  # 现在接收文件ID列表而不是文件名
    
    if not file_ids:
        return jsonify({"error": "No files provided"}), 400
    
    # 检查用户是否有权限操作此项目
    with db.cursor() as cur:
        if user_role == 'LM':
            # LM可以操作任何项目
            cur.execute("SELECT id FROM projectname WHERE id = %s", (project_id,))
        else:
            # BO只能操作自己的项目
            cur.execute("SELECT id FROM projectname WHERE id = %s AND created_by = %s", (project_id, user_id))
        
        project = cur.fetchone()
        if not project:
            return jsonify({"error": "Project not found or you don't have permission"}), 404
        
        # 检查文件是否存在且属于当前用户
        file_ids_str = ','.join(['%s'] * len(file_ids))
        cur.execute(f"""
            SELECT id FROM files 
            WHERE id IN ({file_ids_str}) AND uploaded_by = %s AND isDeleted = FALSE
        """, (*file_ids, user_id))
        found_files = cur.fetchall()
        
        if len(found_files) != len(file_ids):
            return jsonify({"error": "Some files not found or you don't have permission"}), 400
    
    file_type = data.get('fileType')
    notes = data.get('notes', '')
    
    # 创建项目文件记录
    with db.cursor() as cur:
        # 1. 创建项目文件主记录
        sql = """
        INSERT INTO project_files (
            projectId, fileType, notes, files, uploadTime, created_by
        ) VALUES (%s, %s, %s, %s, %s, %s)
        """
        cur.execute(sql, (
            project_id, file_type, notes, ','.join([str(id) for id in file_ids]), datetime.now(), user_id
        ))
        db.commit()
        project_file_id = cur.lastrowid
        
        # 2. 创建项目文件与原始文件的关联
        for file_id in file_ids:
            sql = """
            INSERT INTO project_file_mappings (
                project_file_id, file_id
            ) VALUES (%s, %s)
            """
            cur.execute(sql, (project_file_id, file_id))
        
        db.commit()
    
    return jsonify({"id": project_file_id, "message": "Project files uploaded successfully"}), 201

# 获取项目文件接口
@app.route('/api/project-files/<int:project_id>', methods=['GET'])
@token_required
def get_project_files(project_id):
    # 获取当前用户信息
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    logger.info(f"获取项目 {project_id} 的文件，用户ID: {user_id}, 角色: {user_role}")
    
    try:
        # 验证项目存在并且用户有权限访问
        with db.cursor() as cur:
            # 检查用户是否有权限访问此项目
            if user_role == 'LM':
                # LM可以访问任何项目
                cur.execute("SELECT id FROM projectname WHERE id = %s", (project_id,))
            else:
                # BO只能访问自己的项目
                cur.execute("SELECT id FROM projectname WHERE id = %s AND created_by = %s", (project_id, user_id))
            
            project = cur.fetchone()
            
            if not project:
                logger.warning(f"用户 {user_id} 尝试访问不存在或无权限的项目 {project_id}")
                return jsonify({"error": "Project not found or you don't have permission"}), 404
            
            logger.info(f"用户 {user_id} 有权限访问项目 {project_id}")
            
            # 检查project_files表结构
            try:
                cur.execute("SHOW COLUMNS FROM project_files LIKE 'files'")
                has_files_column = cur.fetchone() is not None
                logger.info(f"project_files表是否有files列: {has_files_column}")
            except Exception as e:
                logger.error(f"检查project_files表结构时出错: {str(e)}")
                has_files_column = False
            
            # 获取项目文件组
            query = """
                SELECT 
                    pf.id, pf.uploadTime as created_at, pf.notes as description, pf.fileType
                FROM project_files pf
                WHERE pf.projectId = %s
            """
            logger.info(f"执行查询项目文件组: {query.replace('%s', str(project_id))}")
            cur.execute(query, (project_id,))
            
            project_files = cur.fetchall()
            
            logger.info(f"找到 {len(project_files)} 个项目文件组")
            
            result = []
            for pf in project_files:
                project_file_id = pf['id']
                
                logger.info(f"处理项目文件组 ID: {project_file_id}, 描述: {pf['description']}")
                
                # 获取关联的原始文件
                query = """
                    SELECT f.id, f.filename, f.originalName, f.filePath as file_path, 
                           f.fileType as file_type, f.uploadTime as created_at,
                           f.uploaded_by, pfm.project_file_id
                    FROM files f
                    JOIN project_file_mappings pfm ON f.id = pfm.file_id
                    WHERE pfm.project_file_id = %s AND f.isDeleted = FALSE
                """
                logger.info(f"执行查询关联文件: {query.replace('%s', str(project_file_id))}")
                cur.execute(query, (project_file_id,))
                
                files = cur.fetchall()
                logger.info(f"找到 {len(files)} 个关联文件")
                
                # 如果没有找到关联文件通过映射表，但表中有files字段，尝试通过files字段获取文件信息
                if len(files) == 0 and has_files_column:
                    logger.info(f"映射表中未找到文件，尝试通过files字段获取")
                    cur.execute("SELECT files FROM project_files WHERE id = %s", (project_file_id,))
                    files_info = cur.fetchone()
                    if files_info and files_info.get('files'):
                        logger.info(f"从files字段获取的文件信息: {files_info['files']}")
                        # 将文件名拆分为列表
                        file_names = [name.strip() for name in files_info['files'].split(',') if name.strip()]
                        if file_names:
                            # 根据文件名查询文件记录
                            names_query = ','.join(['%s'] * len(file_names))
                            query = f"""
                                SELECT id, filename, originalName, filePath as file_path, 
                                       fileType as file_type, uploadTime as created_at,
                                       uploaded_by
                                FROM files
                                WHERE originalName IN ({names_query}) AND isDeleted = FALSE
                            """
                            logger.info(f"根据文件名查询文件: {query}")
                            cur.execute(query, tuple(file_names))
                            files = cur.fetchall()
                            logger.info(f"通过文件名找到 {len(files)} 个文件")
                
                file_list = []
                for file in files:
                    logger.debug(f"处理文件 ID: {file['id']}, 文件名: {file['filename']}, 原始名称: {file['originalName']}")
                    
                    # 构建文件URL
                    file_url = f"/api/files/{file['filename']}"
                    
                    file_list.append({
                        "id": file['id'],
                        "name": file['originalName'] or file['filename'],  # 使用originalName作为显示名称
                        "filename": file['filename'],  # 保留filename字段
                        "url": file_url,
                        "type": file['file_type'],
                        "created_at": file['created_at']
                    })
                
                result.append({
                    "id": pf['id'],
                    "created_at": pf['created_at'],
                    "fileType": pf['fileType'],
                    "description": pf['description'] or f"项目文件 {pf['id']}",
                    "fileList": file_list
                })
            
            logger.info(f"成功获取项目 {project_id} 的文件，共 {len(result)} 个文件组，总计 {sum(len(g['fileList']) for g in result)} 个文件")
            return jsonify(result)
    
    except Exception as e:
        logger.error(f"获取项目文件时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": f"获取项目文件失败: {str(e)}"}), 500

# 邮件接口
@app.route('/api/emails', methods=['POST'])
@token_required
def send_email():
    data = request.json
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    project_id = data.get('projectId')
    
    # 检查用户是否有权限操作此项目
    with db.cursor() as cur:
        if user_role == 'LM':
            # LM可以操作任何项目
            cur.execute("SELECT id FROM projectname WHERE id = %s", (project_id,))
        else:
            # BO只能操作自己的项目
            cur.execute("SELECT id FROM projectname WHERE id = %s AND created_by = %s", (project_id, user_id))
        
        project = cur.fetchone()
        if not project:
            return jsonify({"error": "Project not found or you don't have permission"}), 404
    
    to = data.get('to')
    cc = data.get('cc', '')
    subject = data.get('subject')
    content = data.get('content')
    attachment_ids = data.get('attachmentIds', [])
    
    # 检查文件是否存在并属于当前用户
    if attachment_ids:
        with db.cursor() as cur:
            attachment_ids_str = ','.join(['%s'] * len(attachment_ids))
            cur.execute(f"""
                SELECT id FROM files 
                WHERE id IN ({attachment_ids_str}) AND isDeleted = FALSE
            """, (*attachment_ids,))
            found_files = cur.fetchall()
            
            if len(found_files) != len(attachment_ids):
                return jsonify({"error": "Some attachment files not found"}), 400
    
    # 记录邮件
    with db.cursor() as cur:
        sql = """
        INSERT INTO emails (
            projectId, toRecipient, ccRecipient, subject, content, sendTime, sent_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cur.execute(sql, (
            project_id, to, cc, subject, content, datetime.now(), user_id
        ))
        db.commit()
        email_id = cur.lastrowid
        
        # 关联附件
        if attachment_ids:
            for file_id in attachment_ids:
                cur.execute("""
                    INSERT INTO email_attachments (email_id, file_id)
                    VALUES (%s, %s)
                """, (email_id, file_id))
            db.commit()
    
    # 这里可以添加实际发送邮件的代码
    # 例如使用 smtplib 发送邮件
    
    return jsonify({"id": email_id, "message": "Email sent successfully"}), 200

# 报价接口
@app.route('/api/quotes', methods=['GET'])
@token_required
def get_quotes():
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    # 检查是否有项目ID过滤
    project_id = request.args.get('projectId')
    
    with db.cursor() as cur:
        if project_id:
            # 获取特定项目的任务报价
            if user_role in ['LM', 'FT']:
                # LM和FT可以查看所有报价
                cur.execute("""
                    SELECT * FROM task_quotes 
                    WHERE projectId = %s
                """, (project_id,))
            else:
                # BO只能查看自己项目的报价
                cur.execute("""
                    SELECT tq.* 
                    FROM task_quotes tq
                    JOIN projectname p ON tq.projectId = p.id
                    WHERE tq.projectId = %s AND p.created_by = %s
                """, (project_id, user_id))
        else:
            # 获取所有报价（使用原始quotes表，向后兼容）
            if user_role == 'LM':
                # LM可以查看所有报价
                cur.execute("""
                    SELECT q.*, p.projectName 
                    FROM quotes q 
                    JOIN projectname p ON q.projectId = p.id
                """)
            else:
                # BO只能查看自己项目的报价
                cur.execute("""
                    SELECT q.*, p.projectName 
                    FROM quotes q 
                    JOIN projectname p ON q.projectId = p.id
                    WHERE p.created_by = %s
                """, (user_id,))
        
        quotes = cur.fetchall()
    return jsonify(quotes)

@app.route('/api/quotes', methods=['POST'])
@token_required
def create_quote():
    data = request.json
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    # 只有LM和FT可以创建报价
    if user_role not in ['LM', 'FT']:
        return jsonify({"error": "Only LM or FT can create quotes"}), 403
    
    # 检查请求体中是否包含任务类型 - 判断是创建普通报价还是任务报价
    if 'task' in data:
        # 创建任务报价 (task_quotes表)
        project_id = data.get('projectId')
        task = data.get('task')
        assignee = data.get('assignee')
        language = data.get('language')
        quote_amount = data.get('quoteAmount')
        currency = data.get('currency', 'USD')
        word_count = data.get('wordCount', 0)
        unit_price = data.get('unitPrice')
        deadline = data.get('deadline')
        notes = data.get('notes', '')
        file_id = data.get('fileId')
        status = data.get('status', 'pending')
        extracted_info = data.get('extractedInfo')
        
        # 创建任务报价记录
        with db.cursor() as cur:
            # 先检查task_quotes表是否有extractedInfo列
            try:
                cur.execute("SHOW COLUMNS FROM task_quotes LIKE 'extractedInfo'")
                has_extracted_info_column = cur.fetchone() is not None
                
                if not has_extracted_info_column:
                    # 添加extractedInfo列
                    logger.info("向task_quotes表添加extractedInfo列...")
                    cur.execute("ALTER TABLE task_quotes ADD COLUMN extractedInfo TEXT")
                    db.commit()
                    logger.info("成功添加extractedInfo列到task_quotes表")
            except Exception as e:
                logger.error(f"检查或添加extractedInfo列时出错: {e}")
                # 继续执行，只是不使用extractedInfo字段
            
            # 再次检查extractedInfo列是否存在
            try:
                cur.execute("SHOW COLUMNS FROM task_quotes LIKE 'extractedInfo'")
                has_extracted_info_column = cur.fetchone() is not None
                
                if has_extracted_info_column:
                    # 如果列存在，使用包含extractedInfo的SQL
                    sql = """
                    INSERT INTO task_quotes (
                        projectId, task, assignee, language, quoteAmount, 
                        currency, wordCount, unitPrice, deadline, notes, 
                        fileId, status, createTime, created_by, extractedInfo
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    cur.execute(sql, (
                        project_id, task, assignee, language, quote_amount,
                        currency, word_count, unit_price, deadline, notes,
                        file_id, status, datetime.now(), user_id, extracted_info
                    ))
                else:
                    # 如果列不存在，使用不包含extractedInfo的SQL
                    sql = """
                    INSERT INTO task_quotes (
                        projectId, task, assignee, language, quoteAmount, 
                        currency, wordCount, unitPrice, deadline, notes, 
                        fileId, status, createTime, created_by
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    cur.execute(sql, (
                        project_id, task, assignee, language, quote_amount,
                        currency, word_count, unit_price, deadline, notes,
                        file_id, status, datetime.now(), user_id
                    ))
                
                db.commit()
                quote_id = cur.lastrowid
                
                logger.info(f"任务报价创建成功，ID: {quote_id}")
            except Exception as e:
                logger.error(f"创建任务报价时出错: {e}")
                db.rollback()
                return jsonify({"error": f"Failed to create task quote: {str(e)}"}), 500
        
        return jsonify({"id": quote_id, "message": "Task quote created successfully"}), 201
    else:
        # 创建普通报价 (quotes表) - 保持向后兼容
        project_id = data.get('projectId')
        lsp_name = data.get('lspName')
        quote_amount = data.get('quoteAmount')
        currency = data.get('currency', 'USD')
        word_count = data.get('wordCount', 0)
        quote_date = data.get('quoteDate', datetime.now().strftime('%Y-%m-%d'))
        status = data.get('status', 'pending')
        notes = data.get('notes', '')
        
        # 创建报价记录
        with db.cursor() as cur:
            sql = """
            INSERT INTO quotes (
                projectId, lspName, quoteAmount, currency, wordCount, 
                quoteDate, status, notes, createTime, created_by
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cur.execute(sql, (
                project_id, lsp_name, quote_amount, currency, word_count,
                quote_date, status, notes, datetime.now(), user_id
            ))
            db.commit()
            quote_id = cur.lastrowid
        
        return jsonify({"id": quote_id, "message": "Quote created successfully"}), 201

@app.route('/api/quotes/extract', methods=['POST'])
@token_required
def extract_quote():
    user_role = request.user.get('role')
    
    # 只有LM可以提取报价
    if user_role != 'LM':
        return jsonify({"error": "Only LM can extract quotes"}), 403
    
    data = request.json
    email_content = data.get('emailContent', '')
    
    # 使用正则表达式提取报价信息
    # 这里是一个简单的示例，实际应用中可能需要更复杂的逻辑
    project_id_match = re.search(r'Project ID: (\d+)', email_content)
    amount_match = re.search(r'Quote Amount: (\d+(\.\d+)?)', email_content)
    currency_match = re.search(r'Currency: (\w+)', email_content)
    word_count_match = re.search(r'Word Count: (\d+)', email_content)
    
    extracted_data = {}
    
    if project_id_match:
        extracted_data['projectId'] = int(project_id_match.group(1))
    
    if amount_match:
        extracted_data['quoteAmount'] = float(amount_match.group(1))
    
    if currency_match:
        extracted_data['currency'] = currency_match.group(1)
    
    if word_count_match:
        extracted_data['wordCount'] = int(word_count_match.group(1))
    
    return jsonify(extracted_data)

# 聊天接口
@app.route('/api/chat', methods=['POST'])
@token_required
def chat():
    data = request.json
    user_message = data.get('message')
    response = get_chat_response(user_message)
    return jsonify({'response': response})

# 获取单个邮件及其附件信息
@app.route('/api/emails/<int:email_id>', methods=['GET'])
@token_required
def get_email(email_id):
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    # 检查邮件是否存在并获取邮件信息
    with db.cursor() as cur:
        if user_role == 'LM':
            # LM可以查看任何邮件
            cur.execute("""
                SELECT e.*, p.projectName 
                FROM emails e
                JOIN projectname p ON e.projectId = p.id
                WHERE e.id = %s
            """, (email_id,))
        else:
            # BO只能查看与自己相关的项目的邮件
            cur.execute("""
                SELECT e.*, p.projectName 
                FROM emails e
                JOIN projectname p ON e.projectId = p.id
                WHERE e.id = %s AND p.created_by = %s
            """, (email_id, user_id))
        
        email_info = cur.fetchone()
        if not email_info:
            return jsonify({"error": "Email not found or you don't have permission"}), 404
        
        # 获取邮件附件
        cur.execute("""
            SELECT f.id, f.filename, f.originalName, f.fileType, f.fileSize
            FROM email_attachments ea
            JOIN files f ON ea.file_id = f.id
            WHERE ea.email_id = %s AND f.isDeleted = FALSE
        """, (email_id,))
        
        attachments = []
        for row in cur.fetchall():
            attachments.append({
                'id': row['id'],
                'filename': row['filename'],
                'originalName': row['originalName'],
                'fileType': row['fileType'],
                'fileSize': row['fileSize'],
                'url': f"/api/files/{row['filename']}"
            })
        
        # 将附件信息添加到邮件对象
        email_info['attachments'] = attachments
    
    return jsonify(email_info)

# 获取项目所有邮件
@app.route('/api/project-emails/<int:project_id>', methods=['GET'])
@token_required
def get_project_emails(project_id):
    user_role = request.user.get('role')
    user_id = request.user.get('user_id')
    
    # 检查用户是否有权限访问此项目
    with db.cursor() as cur:
        if user_role == 'LM':
            # LM可以访问任何项目
            cur.execute("SELECT id FROM projectname WHERE id = %s", (project_id,))
        else:
            # BO只能访问自己的项目
            cur.execute("SELECT id FROM projectname WHERE id = %s AND created_by = %s", (project_id, user_id))
        
        project = cur.fetchone()
        if not project:
            return jsonify({"error": "Project not found or you don't have permission"}), 404
    
    # 获取项目相关的邮件
    with db.cursor() as cur:
        cur.execute("""
            SELECT id, projectId, toRecipient, ccRecipient, subject, content, sendTime, sent_by
            FROM emails
            WHERE projectId = %s
            ORDER BY sendTime DESC
        """, (project_id,))
        
        emails = []
        for row in cur.fetchall():
            email_item = dict(row)
            
            # 获取邮件附件数量
            cur.execute("""
                SELECT COUNT(*) as attachmentCount
                FROM email_attachments
                WHERE email_id = %s
            """, (row['id'],))
            attachment_count = cur.fetchone()['attachmentCount']
            email_item['attachmentCount'] = attachment_count
            
            emails.append(email_item)
    
    return jsonify(emails)

# 批量下载文件
@app.route('/api/files/download-batch', methods=['POST'])
@token_required
def download_batch_files():
    data = request.json
    user_id = request.user.get('user_id')
    user_role = request.user.get('role')
    
    file_ids = data.get('fileIds', [])
    if not file_ids:
        return jsonify({"error": "No file IDs provided"}), 400
    
    # 检查文件是否存在并获取文件信息
    with db.cursor() as cur:
        file_ids_str = ','.join(['%s'] * len(file_ids))
        
        if user_role == 'LM':
            # LM可以访问所有文件
            query = f"""
                SELECT id, filename, originalName, filePath 
                FROM files 
                WHERE id IN ({file_ids_str}) AND isDeleted = FALSE
            """
            cur.execute(query, tuple(file_ids))
        else:
            # BO只能访问自己上传的或有权限的文件
            query = f"""
                SELECT id, filename, originalName, filePath 
                FROM files 
                WHERE id IN ({file_ids_str}) AND (uploaded_by = %s OR id IN (
                    SELECT file_id FROM file_permissions WHERE user_id = %s AND can_view = TRUE
                ) OR id IN (
                    SELECT pfm.file_id 
                    FROM project_file_mappings pfm
                    JOIN project_files pf ON pfm.project_file_id = pf.id
                    JOIN projectname p ON pf.projectId = p.id
                    WHERE p.created_by = %s
                )) AND isDeleted = FALSE
            """
            cur.execute(query, (*file_ids, user_id, user_id, user_id))
        
        files_info = cur.fetchall()
        
        if not files_info:
            return jsonify({"error": "No accessible files found"}), 404
    
    # 创建一个临时目录用于存放ZIP文件
    temp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(temp_dir, 'files.zip')
    
    try:
        # 创建ZIP文件
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file_info in files_info:
                file_path = file_info['filePath']
                original_name = file_info['originalName']
                
                # 添加文件到ZIP，使用原始文件名
                if os.path.exists(file_path):
                    zipf.write(file_path, arcname=original_name)
                else:
                    print(f"文件不存在: {file_path}")
        
        # 发送ZIP文件
        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name='files.zip'
        )
    except Exception as e:
        print(f"创建ZIP文件时出错: {str(e)}")
        return jsonify({"error": f"Error creating ZIP file: {str(e)}"}), 500
    finally:
        # 清理临时目录（在请求完成后）
        @after_this_request
        def remove_temp_dir(response):
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                print(f"清理临时目录时出错: {str(e)}")
            return response

# 软删除文件
@app.route('/api/files/<int:file_id>', methods=['DELETE'])
@token_required
def delete_file(file_id):
    user_id = request.user.get('user_id')
    user_role = request.user.get('role')
    
    # 检查文件是否存在
    with db.cursor() as cur:
        cur.execute("SELECT * FROM files WHERE id = %s AND isDeleted = FALSE", (file_id,))
        file_info = cur.fetchone()
        
        if not file_info:
            return jsonify({"error": "File not found"}), 404
        
        # 检查删除权限
        has_permission = False
        
        # LM 可以删除所有文件
        if user_role == 'LM':
            has_permission = True
        # 上传者可以删除自己的文件
        elif file_info['uploaded_by'] == user_id:
            has_permission = True
        else:
            # 检查是否有显式删除权限
            cur.execute("""
                SELECT * FROM file_permissions 
                WHERE file_id = %s AND user_id = %s AND can_delete = TRUE
            """, (file_id, user_id))
            permission = cur.fetchone()
            has_permission = permission is not None
        
        if not has_permission:
            return jsonify({"error": "You don't have permission to delete this file"}), 403
        
        # 软删除文件（将isDeleted标记为TRUE）
        cur.execute("UPDATE files SET isDeleted = TRUE WHERE id = %s", (file_id,))
        db.commit()

    return jsonify({"message": "File deleted successfully"}), 200

# 批量软删除文件
@app.route('/api/files/delete-batch', methods=['POST'])
@token_required
def delete_batch_files():
    data = request.json
    user_id = request.user.get('user_id')
    user_role = request.user.get('role')
    
    file_ids = data.get('fileIds', [])
    if not file_ids:
        return jsonify({"error": "No file IDs provided"}), 400
    
    # 获取有权限删除的文件
    with db.cursor() as cur:
        deleted_count = 0
        
        for file_id in file_ids:
            # 检查文件是否存在
            cur.execute("SELECT * FROM files WHERE id = %s AND isDeleted = FALSE", (file_id,))
            file_info = cur.fetchone()
            
            if not file_info:
                continue
            
            # 检查删除权限
            has_permission = False
            
            # LM 可以删除所有文件
            if user_role == 'LM':
                has_permission = True
            # 上传者可以删除自己的文件
            elif file_info['uploaded_by'] == user_id:
                has_permission = True
            else:
                # 检查是否有显式删除权限
                cur.execute("""
                    SELECT * FROM file_permissions 
                    WHERE file_id = %s AND user_id = %s AND can_delete = TRUE
                """, (file_id, user_id))
                permission = cur.fetchone()
                has_permission = permission is not None
            
            if has_permission:
                # 软删除文件
                cur.execute("UPDATE files SET isDeleted = TRUE WHERE id = %s", (file_id,))
                deleted_count += 1
        
        db.commit()
    
    return jsonify({"message": f"{deleted_count} files deleted successfully"}), 200

# 添加一个诊断路由，用于检查文件上传和关联情况
@app.route('/api/diagnostics/file-relations', methods=['GET'])
@token_required
def diagnose_file_relations():
    user_id = request.user.get('user_id')
    user_role = request.user.get('role')
    
    try:
        with db.cursor() as cur:
            # 获取用户上传的文件数量
            cur.execute("SELECT COUNT(*) as count FROM files WHERE uploaded_by = %s AND isDeleted = FALSE", (user_id,))
            file_count = cur.fetchone()['count']
            
            # 获取用户参与的项目数量
            if user_role == 'LM':
                cur.execute("SELECT COUNT(*) as count FROM projectname")
            else:
                cur.execute("SELECT COUNT(*) as count FROM projectname WHERE created_by = %s", (user_id,))
            project_count = cur.fetchone()['count']
            
            # 获取项目文件数量
            if user_role == 'LM':
                cur.execute("SELECT COUNT(*) as count FROM project_files")
            else:
                cur.execute("SELECT COUNT(*) as count FROM project_files WHERE created_by = %s", (user_id,))
            project_file_count = cur.fetchone()['count']
            
            # 获取文件映射数量
            cur.execute("""
                SELECT COUNT(*) as count FROM project_file_mappings pfm
                JOIN project_files pf ON pfm.project_file_id = pf.id
                JOIN files f ON pfm.file_id = f.id
                WHERE f.uploaded_by = %s AND f.isDeleted = FALSE
            """, (user_id,))
            mapping_count = cur.fetchone()['count']
            
            # 获取最近上传的文件
            cur.execute("""
                SELECT id, filename, originalName, uploadTime, isDeleted
                FROM files
                WHERE uploaded_by = %s
                ORDER BY uploadTime DESC
                LIMIT 5
            """, (user_id,))
            recent_files = cur.fetchall()
            
            # 获取最近创建的项目文件
            if user_role == 'LM':
                cur.execute("""
                    SELECT pf.id, pf.projectId, pf.fileType, pf.uploadTime, pf.files
                    FROM project_files pf
                    ORDER BY pf.uploadTime DESC
                    LIMIT 5
                """)
            else:
                cur.execute("""
                    SELECT pf.id, pf.projectId, pf.fileType, pf.uploadTime, pf.files
                    FROM project_files pf
                    WHERE pf.created_by = %s
                    ORDER BY pf.uploadTime DESC
                    LIMIT 5
                """, (user_id,))
            recent_project_files = cur.fetchall()
            
            # 获取文件映射详情
            cur.execute("""
                SELECT pfm.id, pfm.project_file_id, pfm.file_id,
                       pf.projectId, f.originalName
                FROM project_file_mappings pfm
                JOIN project_files pf ON pfm.project_file_id = pf.id
                JOIN files f ON pfm.file_id = f.id
                WHERE f.uploaded_by = %s AND f.isDeleted = FALSE
                ORDER BY pfm.id DESC
                LIMIT 10
            """, (user_id,))
            mapping_details = cur.fetchall()
            
            # 检查数据库表结构
            cur.execute("SHOW COLUMNS FROM project_files")
            project_files_columns = [col['Field'] for col in cur.fetchall()]
            
            cur.execute("SHOW COLUMNS FROM project_file_mappings")
            mapping_columns = [col['Field'] for col in cur.fetchall()]
            
            # 返回诊断结果
            return jsonify({
                "user_id": user_id,
                "user_role": user_role,
                "file_count": file_count,
                "project_count": project_count,
                "project_file_count": project_file_count,
                "mapping_count": mapping_count,
                "recent_files": recent_files,
                "recent_project_files": recent_project_files,
                "mapping_details": mapping_details,
                "project_files_columns": project_files_columns,
                "mapping_columns": mapping_columns,
                "database_check": "ok"
            })
            
    except Exception as e:
        logger.error(f"诊断文件关联时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            "error": str(e),
            "traceback": traceback.format_exc()
        }), 500

# 初始化数据库函数
def init_db():
    logger.info("开始初始化数据库表...")
    
    with db.cursor() as cur:
        # 检查users表是否存在
        cur.execute("SHOW TABLES LIKE 'users'")
        if not cur.fetchone():
            # 创建users表
            logger.info("创建users表...")
            cur.execute("""
                CREATE TABLE users (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    username VARCHAR(50) NOT NULL UNIQUE,
                    password VARCHAR(100) NOT NULL,
                    name VARCHAR(100) NOT NULL,
                    role ENUM('LM', 'BO', 'FT') NOT NULL,
                    email VARCHAR(100),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # 添加默认用户
            logger.info("添加默认用户...")
            cur.execute("""
                INSERT INTO users (username, password, name, role, email)
                VALUES 
                ('admin', 'admin123', 'Admin User', 'LM', 'admin@example.com'),
                ('bo1', 'bo123', 'Business Owner 1', 'BO', 'bo1@example.com'),
                ('bo2', 'bo123', 'Business Owner 2', 'BO', 'bo2@example.com'),
                ('FT', 'ft123', 'Financial Team', 'FT', 'finance@example.com')
            """)
            
            logger.info("Users表创建完成，添加了默认用户")
        else:
            logger.info("Users表已存在，无需创建")
            
            # 检查users表的role字段是否包含FT角色
            cur.execute("SHOW COLUMNS FROM users LIKE 'role'")
            role_col = cur.fetchone()
            if role_col and 'FT' not in role_col['Type']:
                logger.info("修改users表role字段以支持FT角色...")
                cur.execute("ALTER TABLE users MODIFY COLUMN role ENUM('LM', 'BO', 'FT') NOT NULL")
                logger.info("users表role字段已更新")
            
            # 检查是否已存在FT用户
            cur.execute("SELECT id FROM users WHERE username = 'FT'")
            if not cur.fetchone():
                logger.info("添加财务团队默认用户...")
                cur.execute("""
                    INSERT INTO users (username, password, name, role, email)
                    VALUES ('FT', 'ft123', 'Financial Team', 'FT', 'finance@example.com')
                """)
                logger.info("财务团队用户已添加")
        
        # 检查projectname表是否有created_by字段
        cur.execute("SHOW COLUMNS FROM projectname LIKE 'created_by'")
        if not cur.fetchone():
            # 添加created_by字段
            logger.info("向projectname表添加created_by字段...")
            cur.execute("ALTER TABLE projectname ADD COLUMN created_by INT")
            logger.info("成功添加created_by字段到projectname表")
        else:
            logger.info("projectname表已有created_by字段")
        
        # 检查requests表是否有created_by字段
        cur.execute("SHOW COLUMNS FROM requests LIKE 'created_by'")
        if not cur.fetchone():
            # 添加created_by字段
            logger.info("向requests表添加created_by字段...")
            cur.execute("ALTER TABLE requests ADD COLUMN created_by INT")
            logger.info("成功添加created_by字段到requests表")
        else:
            logger.info("requests表已有created_by字段")
        
        # 检查requests表是否有project_id字段
        cur.execute("SHOW COLUMNS FROM requests LIKE 'project_id'")
        if not cur.fetchone():
            # 添加project_id字段
            logger.info("向requests表添加project_id字段...")
            cur.execute("ALTER TABLE requests ADD COLUMN project_id INT")
            logger.info("成功添加project_id字段到requests表")
        else:
            logger.info("requests表已有project_id字段")
        
        # 检查files表是否存在
        cur.execute("SHOW TABLES LIKE 'files'")
        if not cur.fetchone():
            # 创建files表
            logger.info("创建files表...")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS files (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    filename VARCHAR(255) NOT NULL,
                    originalName VARCHAR(255) NOT NULL,
                    fileType VARCHAR(100),
                    fileSize INT,
                    filePath VARCHAR(255) NOT NULL,
                    fileContent LONGBLOB,
                    uploadTime DATETIME NOT NULL,
                    uploaded_by INT NOT NULL,
                    isDeleted BOOLEAN DEFAULT FALSE,
                    FOREIGN KEY (uploaded_by) REFERENCES users(id)
                )
            """)
            logger.info("Files表创建完成")
        else:
            logger.info("Files表已存在，无需创建")
        
        # 检查task_quotes表是否存在
        cur.execute("SHOW TABLES LIKE 'task_quotes'")
        if not cur.fetchone():
            logger.info("创建task_quotes表...")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS task_quotes (
                  id INT AUTO_INCREMENT PRIMARY KEY,
                  projectId INT NOT NULL,
                  task ENUM('translation', 'lqa', 'translationUpdate', 'lqaReportFinalization') NOT NULL,
                  assignee VARCHAR(100) NOT NULL,
                  language VARCHAR(10) NOT NULL,
                  quoteAmount DECIMAL(10, 2) NOT NULL,
                  currency VARCHAR(3) NOT NULL DEFAULT 'USD',
                  wordCount INT NOT NULL DEFAULT 0,
                  unitPrice DECIMAL(10, 4),
                  deadline DATE,
                  notes TEXT,
                  fileId INT,
                  extractedInfo TEXT,
                  status ENUM('pending', 'approved', 'rejected') NOT NULL DEFAULT 'pending',
                  createTime DATETIME NOT NULL,
                  created_by INT NOT NULL,
                  FOREIGN KEY (projectId) REFERENCES projectname(id) ON DELETE CASCADE,
                  FOREIGN KEY (fileId) REFERENCES files(id) ON DELETE SET NULL,
                  FOREIGN KEY (created_by) REFERENCES users(id)
                )
            """)
            
            # 创建索引
            logger.info("为task_quotes表创建索引...")
            cur.execute("CREATE INDEX idx_task_quotes_project ON task_quotes(projectId)")
            cur.execute("CREATE INDEX idx_task_quotes_language ON task_quotes(language)")
            cur.execute("CREATE INDEX idx_task_quotes_task ON task_quotes(task)")
            cur.execute("CREATE INDEX idx_task_quotes_status ON task_quotes(status)")
            
            logger.info("task_quotes表创建完成")
        else:
            logger.info("task_quotes表已存在，无需创建")
            
            # 检查是否存在extractedInfo列
            try:
                cur.execute("SHOW COLUMNS FROM task_quotes LIKE 'extractedInfo'")
                has_extracted_info = cur.fetchone() is not None
                if not has_extracted_info:
                    logger.info("向现有task_quotes表添加extractedInfo列")
                    cur.execute("ALTER TABLE task_quotes ADD COLUMN extractedInfo TEXT")
                    db.commit()
                    logger.info("成功添加extractedInfo列到现有task_quotes表")
            except Exception as e:
                logger.error(f"检查或添加extractedInfo列时出错: {e}")
                # 继续执行，不阻止其他操作
        
        # 创建视图，用于查询报价详情
        cur.execute("SHOW TABLES LIKE 'task_quotes_view'")
        if not cur.fetchone():
            logger.info("创建task_quotes_view视图...")
            cur.execute("""
                CREATE OR REPLACE VIEW task_quotes_view AS
                SELECT 
                  tq.*,
                  p.projectName,
                  p.projectStatus,
                  p.requestName,
                  p.projectManager,
                  f.originalName as fileName
                FROM task_quotes tq
                JOIN projectname p ON tq.projectId = p.id
                LEFT JOIN files f ON tq.fileId = f.id
            """)
            logger.info("task_quotes_view视图创建完成")
        else:
            logger.info("task_quotes_view视图已存在，无需创建")
        
        db.commit()
    
    logger.info("数据库初始化完成")

@app.route('/api/project-task-assignments/<int:project_id>', methods=['GET'])
@token_required
def get_project_task_assignments(project_id):
    user_role = request.user.get('role')
    user_id = request.user.get('user_id') or request.user.get('id')
    
    try:
        # 检查项目是否存在
        with db.cursor() as cur:
            cur.execute("SELECT id FROM projectname WHERE id = %s", (project_id,))
            project = cur.fetchone()
            if not project:
                return jsonify({"error": "Project not found"}), 404
            
            # 检查表是否存在
            cur.execute("SHOW TABLES LIKE 'project_task_assignments'")
            table_exists = cur.fetchone() is not None
            
            if not table_exists:
                # 表不存在，返回空数组
                return jsonify([])
            
            # 获取所有任务分配
            cur.execute("""
                SELECT id, project_id, task_type, language, assignee, deadline, notes, createTime
                FROM project_task_assignments
                WHERE project_id = %s
                ORDER BY task_type, language
            """, (project_id,))
            
            assignments = cur.fetchall()
            
            # 处理日期字段
            for assignment in assignments:
                if 'deadline' in assignment and assignment['deadline']:
                    assignment['deadline'] = assignment['deadline'].isoformat()
                if 'createTime' in assignment and assignment['createTime']:
                    assignment['createTime'] = assignment['createTime'].isoformat()
            
            return jsonify(assignments)
    except Exception as e:
        print(f"获取项目任务分配时出错: {e}")
        return jsonify({"error": f"Failed to get project task assignments: {str(e)}"}), 500

# 添加导出报价信息的路由
@app.route('/api/quotes/export', methods=['GET'])
@token_required
def export_quotes():
    try:
        # 获取当前用户，确保是FT或LM角色
        # 从token_required装饰器中获取用户信息，而不是使用get_jwt_identity
        user_role = request.user.get('role')
        
        if user_role not in ['FT', 'LM', 'PM']:
            return jsonify({'error': 'Permission denied: Only Financial Team, Localization Manager or Project Manager can export quotes'}), 403
        
        # 获取项目ID
        project_id = request.args.get('projectId')
        if not project_id:
            return jsonify({'error': 'Missing project ID'}), 400
        
        # 连接数据库
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 获取项目信息
        cursor.execute('''
            SELECT * FROM projectname WHERE id = %s
        ''', (project_id,))
        project = cursor.fetchone()
        
        if not project:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Project not found'}), 404
        
        # 获取项目任务分配情况
        cursor.execute('''
            SELECT * FROM projectname_task_assignments 
            WHERE projectId = %s
        ''', (project_id,))
        task_assignments = cursor.fetchall()
        
        # 获取项目报价信息
        cursor.execute('''
            SELECT tq.*, f.originalName as fileName 
            FROM task_quotes tq
            LEFT JOIN files f ON tq.fileId = f.id
            WHERE tq.projectId = %s
        ''', (project_id,))
        quotes = cursor.fetchall()
        
        # 创建一个Excel文件
        output = io.BytesIO()
        
        # 创建一个工作簿和一个总览工作表
        workbook = openpyxl.Workbook()
        overview_sheet = workbook.active
        overview_sheet.title = "Project_Overview"
        
        # 设置表头样式
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 添加项目基本信息到总览表
        overview_sheet.append(["Project Information", ""])
        overview_sheet.append(["Project Name", project['projectName']])
        overview_sheet.append(["Project Status", project['projectStatus']])
        overview_sheet.append(["Request Name", project['requestName']])
        overview_sheet.append(["Project Manager", project['projectManager']])
        overview_sheet.append(["Create Time", project['createTime'].strftime('%Y-%m-%d %H:%M:%S') if project['createTime'] else 'N/A'])
        overview_sheet.append(["Source Language", project['sourceLanguage']])
        overview_sheet.append(["Target Languages", project['targetLanguages']])
        overview_sheet.append(["Word Count", project['wordCount']])
        
        # 添加空行
        overview_sheet.append([])
        
        # 添加所有报价信息表头
        overview_sheet.append([
            "Task Type",
            "Assignee",
            "Language",
            "Quote Amount",
            "Currency",
            "Word Count",
            "Unit Price",
            "Deadline",
            "Status",
            "Notes",
            "File Name"
        ])
        
        # 设置表头样式
        for col in range(1, 12):
            cell = overview_sheet.cell(row=11, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 添加所有报价信息数据
        row_num = 12
        for quote in quotes:
            task_type_map = {
                'translation': 'Translation',
                'lqa': 'LQA',
                'translationUpdate': 'Translation Update',
                'lqaReportFinalization': 'LQA Report Finalization'
            }
            task_type = task_type_map.get(quote['task'], quote['task'])
            
            overview_sheet.append([
                task_type,
                quote['assignee'],
                quote['language'],
                quote['quoteAmount'],
                quote['currency'],
                quote['wordCount'],
                quote['unitPrice'],
                quote['deadline'].strftime('%Y-%m-%d') if quote['deadline'] else 'N/A',
                quote['status'],
                quote['notes'],
                quote['fileName'] if quote['fileName'] else 'N/A'
            ])
            
            # 设置数据单元格样式
            for col in range(1, 12):
                cell = overview_sheet.cell(row=row_num, column=col)
                cell.border = thin_border
                
                # 对报价金额和单价进行数字格式化
                if col == 4:  # 报价金额
                    cell.number_format = '#,##0.00'
                elif col == 7:  # 单价
                    cell.number_format = '#,##0.0000'
            
            row_num += 1
        
        # 调整列宽
        for col in overview_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            overview_sheet.column_dimensions[column].width = adjusted_width
        
        # 收集所有语言代码
        languages = set()
        for quote in quotes:
            languages.add(quote['language'])
        
        # 为每种语言创建一个工作表
        for language in languages:
            # 创建语言特定的工作表
            lang_sheet = workbook.create_sheet(title=f"{language}")
            
            # 添加项目基本信息
            lang_sheet.append(["Project Information", ""])
            lang_sheet.append(["Project Name", project['projectName']])
            lang_sheet.append(["Project Status", project['projectStatus']])
            lang_sheet.append(["Language", language])
            lang_sheet.append(["Word Count", project['wordCount']])
            
            lang_sheet.append([])
            
            # 添加该语言的报价信息表头
            lang_sheet.append([
                "Task Type",
                "Assignee",
                "Quote Amount",
                "Currency",
                "Word Count",
                "Unit Price",
                "Deadline",
                "Status",
                "Notes",
                "File Name"
            ])
            
            # 设置表头样式
            for col in range(1, 11):
                cell = lang_sheet.cell(row=7, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 筛选该语言的报价数据
            language_quotes = [q for q in quotes if q['language'] == language]
            
            # 添加该语言的报价信息
            row_num = 8
            for quote in language_quotes:
                task_type_map = {
                    'translation': 'Translation',
                    'lqa': 'LQA',
                    'translationUpdate': 'Translation Update',
                    'lqaReportFinalization': 'LQA Report Finalization'
                }
                task_type = task_type_map.get(quote['task'], quote['task'])
                
                lang_sheet.append([
                    task_type,
                    quote['assignee'],
                    quote['quoteAmount'],
                    quote['currency'],
                    quote['wordCount'],
                    quote['unitPrice'],
                    quote['deadline'].strftime('%Y-%m-%d') if quote['deadline'] else 'N/A',
                    quote['status'],
                    quote['notes'],
                    quote['fileName'] if quote['fileName'] else 'N/A'
                ])
                
                # 设置数据单元格样式
                for col in range(1, 11):
                    cell = lang_sheet.cell(row=row_num, column=col)
                    cell.border = thin_border
                    
                    # 对报价金额和单价进行数字格式化
                    if col == 3:  # 报价金额
                        cell.number_format = '#,##0.00'
                    elif col == 6:  # 单价
                        cell.number_format = '#,##0.0000'
                
                row_num += 1
            
            # 调整列宽
            for col in lang_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                lang_sheet.column_dimensions[column].width = adjusted_width
        
        # 保存工作簿到输出流
        workbook.save(output)
        output.seek(0)
        
        # 释放数据库连接
        cursor.close()
        conn.close()
        
        # 生成文件名
        safe_project_name = secure_filename(project['projectName'])
        filename = f"Project_Quote_{safe_project_name}_{project_id}.xlsx"
        
        # 返回Excel文件
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=filename,
            as_attachment=True
        )
    
    except Exception as e:
        print(f"Error exporting quotes: {str(e)}")
        return jsonify({'error': f'Failed to export quotes: {str(e)}'}), 500

# 添加修复文件映射的API端点
@app.route('/api/fix-file-mappings', methods=['POST'])
@token_required
def fix_file_mappings():
    """修复文件映射关系，解决文件上传后未显示在项目中的问题"""
    try:
        # 提取请求数据
        data = request.json if request.is_json else {}
        user_id = request.user.get('user_id') or request.user.get('id')
        user_role = request.user.get('role')
        
        # 指定要修复的项目ID (如果提供)
        project_id = data.get('projectId')
        
        logger.info(f"开始修复文件映射，用户ID: {user_id}, 角色: {user_role}, 项目ID: {project_id}")
        
        with db.cursor() as cur:
            fixed_count = 0
            
            # 获取用户上传的文件
            if project_id:
                # 如果指定了项目ID，仅修复该项目的文件
                # 首先检查用户是否有权限访问此项目
                if user_role in ['LM', 'FT']:
                    cur.execute("SELECT id FROM projectname WHERE id = %s", (project_id,))
                else:
                    cur.execute("SELECT id FROM projectname WHERE id = %s AND created_by = %s", (project_id, user_id))
                
                project = cur.fetchone()
                if not project:
                    response = jsonify({"error": "Project not found or you don't have permission"})
                    # 添加CORS头
                    response.headers.add('Access-Control-Allow-Origin', '*')
                    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
                    return response, 404
                
                # 获取该用户上传的且没有映射关系的文件
                cur.execute("""
                    SELECT f.id, f.filename, f.originalName, f.uploadTime
                    FROM files f
                    LEFT JOIN project_file_mappings pfm ON f.id = pfm.file_id
                    LEFT JOIN project_files pf ON pfm.project_file_id = pf.id AND pf.projectId = %s
                    WHERE f.uploaded_by = %s AND f.isDeleted = FALSE AND pfm.id IS NULL
                    ORDER BY f.uploadTime DESC
                """, (project_id, user_id))
            else:
                # 没有指定项目ID，获取所有未映射的文件
                cur.execute("""
                    SELECT f.id, f.filename, f.originalName, f.uploadTime
                    FROM files f
                    LEFT JOIN project_file_mappings pfm ON f.id = pfm.file_id
                    WHERE f.uploaded_by = %s AND f.isDeleted = FALSE AND pfm.id IS NULL
                    ORDER BY f.uploadTime DESC
                """, (user_id,))
            
            unmapped_files = cur.fetchall()
            logger.info(f"找到 {len(unmapped_files)} 个未映射的文件")
            
            if not unmapped_files:
                response = jsonify({"message": "No files need to be fixed", "fixed_count": 0})
                # 添加CORS头
                response.headers.add('Access-Control-Allow-Origin', '*')
                response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
                return response
            
            # 如果指定了项目ID，使用该项目
            if project_id:
                # 查找该项目的现有项目文件记录
                cur.execute("""
                    SELECT id FROM project_files 
                    WHERE projectId = %s AND fileType = 'source'
                    ORDER BY uploadTime DESC LIMIT 1
                """, (project_id,))
                
                project_file = cur.fetchone()
                
                if project_file:
                    # 使用现有的项目文件记录
                    project_file_id = project_file['id']
                    logger.info(f"使用现有项目文件记录ID: {project_file_id}")
                else:
                    # 创建新的项目文件记录
                    file_names = [f['originalName'] for f in unmapped_files]
                    files_str = ', '.join(file_names)
                    
                    cur.execute("""
                        INSERT INTO project_files 
                        (projectId, fileType, notes, files, uploadTime, created_by)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (
                        project_id, 
                        'source', 
                        '', 
                        files_str, 
                        datetime.now(), 
                        user_id
                    ))
                    db.commit()
                    project_file_id = cur.lastrowid
                    logger.info(f"创建新项目文件记录ID: {project_file_id}")
                
                # 为每个未映射的文件创建映射
                for file in unmapped_files:
                    try:
                        cur.execute("""
                            INSERT INTO project_file_mappings 
                            (project_file_id, file_id)
                            VALUES (%s, %s)
                        """, (project_file_id, file['id']))
                        fixed_count += 1
                        logger.info(f"创建映射: 项目文件ID {project_file_id} -> 文件ID {file['id']} (文件名: {file['originalName']})")
                    except Exception as e:
                        logger.error(f"创建映射时出错: {str(e)}")
                
                db.commit()
                logger.info(f"已修复 {fixed_count} 个文件映射")
                
                response = jsonify({
                    "message": f"Successfully fixed {fixed_count} file mappings",
                    "fixed_count": fixed_count,
                    "project_id": project_id,
                    "project_file_id": project_file_id
                })
                # 添加CORS头
                response.headers.add('Access-Control-Allow-Origin', '*')
                response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
                return response
            else:
                # 未指定项目ID，尝试为每个文件找到合适的项目
                # 按上传者分组文件
                files_by_uploader = {}
                for file in unmapped_files:
                    if file['uploaded_by'] not in files_by_uploader:
                        files_by_uploader[file['uploaded_by']] = []
                    files_by_uploader[file['uploaded_by']].append(file)
                
                # 处理每个上传者的文件
                for uploader_id, files_group in files_by_uploader.items():
                    # 查找该用户最近的项目
                    cur.execute("""
                        SELECT id FROM projectname 
                        WHERE created_by = %s 
                        ORDER BY createTime DESC LIMIT 1
                    """, (uploader_id,))
                    
                    user_project = cur.fetchone()
                    if not user_project:
                        logger.warning(f"用户 {uploader_id} 没有关联的项目，跳过其文件")
                        continue
                    
                    user_project_id = user_project['id']
                    logger.info(f"找到用户 {uploader_id} 的最近项目ID: {user_project_id}")
                    
                    # 查找该项目的现有项目文件记录
                    cur.execute("""
                        SELECT id FROM project_files 
                        WHERE projectId = %s AND fileType = 'source'
                        ORDER BY uploadTime DESC LIMIT 1
                    """, (user_project_id,))
                    
                    project_file = cur.fetchone()
                    
                    if project_file:
                        # 使用现有的项目文件记录
                        project_file_id = project_file['id']
                        logger.info(f"使用现有项目文件记录ID: {project_file_id}")
                    else:
                        # 创建新的项目文件记录
                        file_names = [f['originalName'] for f in files_group]
                        files_str = ', '.join(file_names)
                        
                        cur.execute("""
                            INSERT INTO project_files 
                            (projectId, fileType, notes, files, uploadTime, created_by)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            user_project_id, 
                            'source', 
                            '', 
                            files_str, 
                            datetime.now(), 
                            uploader_id
                        ))
                        db.commit()
                        project_file_id = cur.lastrowid
                        logger.info(f"创建新项目文件记录ID: {project_file_id}")
                    
                    # 为每个未映射的文件创建映射
                    for file in files_group:
                        try:
                            cur.execute("""
                                INSERT INTO project_file_mappings 
                                (project_file_id, file_id)
                                VALUES (%s, %s)
                            """, (project_file_id, file['id']))
                            fixed_count += 1
                            logger.info(f"创建映射: 项目文件ID {project_file_id} -> 文件ID {file['id']} (文件名: {file['originalName']})")
                        except Exception as e:
                            logger.error(f"创建映射时出错: {str(e)}")
                    
                    db.commit()
                
                logger.info(f"总共修复了 {fixed_count} 个文件映射")
                
                response = jsonify({
                    "message": f"Successfully fixed {fixed_count} file mappings",
                    "fixed_count": fixed_count
                })
                # 添加CORS头
                response.headers.add('Access-Control-Allow-Origin', '*')
                response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
                return response
    
    except Exception as e:
        logger.error(f"修复文件映射时出错: {str(e)}")
        logger.error(traceback.format_exc())
        response = jsonify({"error": f"Failed to fix file mappings: {str(e)}"})
        # 添加CORS头
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        return response, 500

# 在token_required装饰器上面添加一个预检请求处理器
@app.route('/api/fix-file-mappings', methods=['OPTIONS'])
def handle_fix_file_mappings_preflight():
    response = make_response()
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# 获取项目报价列表
@app.route('/api/projects/<int:project_id>/quotes', methods=['GET'])
@token_required
def get_project_quotes(project_id):
    try:
        # 验证用户权限
        user_role = request.user.get('role')
        if user_role not in ['PM', 'LM', 'FT']:
            return jsonify({'error': 'Permission denied'}), 403
        
        # 从数据库获取项目报价列表
        with db.cursor(dictionary=True) as cur:
            cur.execute("""
                SELECT tq.*, f.originalName as fileName, f.fileSize, f.mimeType
                FROM task_quotes tq
                LEFT JOIN files f ON tq.fileId = f.id
                WHERE tq.projectId = %s
                ORDER BY tq.createTime DESC
            """, (project_id,))
            quotes = cur.fetchall()
            
            # 处理日期格式，使其可JSON序列化
            for quote in quotes:
                if quote.get('deadline') and isinstance(quote['deadline'], date):
                    quote['deadline'] = quote['deadline'].isoformat()
                if quote.get('createTime') and isinstance(quote['createTime'], datetime):
                    quote['createTime'] = quote['createTime'].isoformat()
                
                # 检查是否有提取的特定列数据
                if quote.get('extractedInfo'):
                    # 尝试解析JSON数据，如果失败则保留原始字符串
                    try:
                        quote['hasExtractedInfo'] = True
                    except:
                        quote['hasExtractedInfo'] = True
                else:
                    quote['hasExtractedInfo'] = False
            
            return jsonify(quotes), 200
                
    except Exception as e:
        logger.error(f"Error getting project quotes: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

# 获取特定报价的提取数据
@app.route('/api/quotes/<int:quote_id>/extracted-info', methods=['GET'])
@token_required
def get_quote_extracted_info(quote_id):
    try:
        # 验证用户权限
        user_role = request.user.get('role')
        if user_role not in ['PM', 'LM', 'FT']:
            return jsonify({'error': 'Permission denied'}), 403
        
        # 从数据库获取报价的extractedInfo
        with db.cursor(dictionary=True) as cur:
            cur.execute("""
                SELECT extractedInfo FROM task_quotes WHERE id = %s
            """, (quote_id,))
            result = cur.fetchone()
            
            if not result:
                return jsonify({'error': 'Quote not found'}), 404
            
            extracted_info = result.get('extractedInfo')
            
            # 如果没有提取数据
            if not extracted_info:
                return jsonify({'extractedInfo': []}), 200
            
            # 尝试解析JSON数据
            try:
                extracted_info_json = json.loads(extracted_info)
                return jsonify({'extractedInfo': extracted_info_json}), 200
            except Exception as e:
                # 如果JSON解析失败，返回原始字符串
                return jsonify({'extractedInfo': extracted_info, 'error': 'Failed to parse JSON'}), 200
                
    except Exception as e:
        logger.error(f"Error getting quote extracted info: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

if __name__ == '__main__':
    init_db() 
    app.run(debug=True)